import pyautogui
import keyboard
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os
import time
import pygetwindow as gw
from PIL import ImageGrab
import ctypes
from ctypes import wintypes
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading

# ─── Default Paths ────────────────────────────────────────────────────────────
DEFAULT_EXCEL_FILE = r"C:\Users\zhili\OneDrive\Desktop\Python\tax_records.xlsx"
DEFAULT_TEMP_FOLDER = r"C:\Users\zhili\OneDrive\Desktop\Python"

# Global path variables (updated by GUI)
EXCEL_FILE = DEFAULT_EXCEL_FILE
TEMP_FOLDER = DEFAULT_TEMP_FOLDER

# ─── Taskbar / Window Helpers ──────────────────────────────────────────────────

def get_taskbar_info():
    """Get taskbar position and size"""
    try:
        taskbar_hwnd = ctypes.windll.user32.FindWindowW("Shell_TrayWnd", None)
        if not taskbar_hwnd:
            taskbar_hwnd = ctypes.windll.user32.FindWindowW("Shell_SecondaryTrayWnd", None)

        if taskbar_hwnd:
            rect = wintypes.RECT()
            ctypes.windll.user32.GetWindowRect(taskbar_hwnd, ctypes.byref(rect))

            screen_width = ctypes.windll.user32.GetSystemMetrics(0)
            screen_height = ctypes.windll.user32.GetSystemMetrics(1)

            if rect.top >= screen_height - 100:
                position = "bottom"
            elif rect.left <= 10:
                position = "left"
            elif rect.top <= 10:
                position = "top"
            elif rect.left >= screen_width - 100:
                position = "right"
            else:
                position = "unknown"

            print(f"Taskbar found at {position}: {rect.left},{rect.top} to {rect.right},{rect.bottom}")
            return (rect.left, rect.top, rect.right, rect.bottom), position
    except Exception as e:
        print(f"Error finding taskbar: {e}")

    return None, "unknown"


def get_window_with_taskbar_bounds():
    """Get bounds that include both the active window and taskbar"""
    try:
        hwnd = ctypes.windll.user32.GetForegroundWindow()
        if not hwnd:
            print("No foreground window found")
            return None, "Unknown"

        window_rect = wintypes.RECT()
        DWMWA_EXTENDED_FRAME_BOUNDS = 9
        result = ctypes.windll.dwmapi.DwmGetWindowAttribute(
            hwnd,
            DWMWA_EXTENDED_FRAME_BOUNDS,
            ctypes.byref(window_rect),
            ctypes.sizeof(window_rect)
        )

        if result != 0:
            ctypes.windll.user32.GetWindowRect(hwnd, ctypes.byref(window_rect))

        length = 256
        title_buffer = ctypes.create_unicode_buffer(length)
        ctypes.windll.user32.GetWindowTextW(hwnd, title_buffer, length)
        window_title = title_buffer.value

        taskbar_bounds, taskbar_position = get_taskbar_info()

        screen_width = ctypes.windll.user32.GetSystemMetrics(0)
        screen_height = ctypes.windll.user32.GetSystemMetrics(1)

        left = window_rect.left
        top = window_rect.top
        right = window_rect.right
        bottom = window_rect.bottom

        if taskbar_bounds:
            tb_left, tb_top, tb_right, tb_bottom = taskbar_bounds
            left = min(left, tb_left)
            top = min(top, tb_top)
            right = max(right, tb_right)
            bottom = max(bottom, tb_bottom)

        left = max(0, left)
        top = max(0, top)
        right = min(screen_width, right)
        bottom = min(screen_height, bottom)

        print(f"Window: '{window_title}'")
        print(f"Window bounds: {window_rect.left},{window_rect.top} to {window_rect.right},{window_rect.bottom}")
        print(f"Final bounds (with taskbar): {left},{top} to {right},{bottom}")

        return (int(left), int(top), int(right), int(bottom)), window_title

    except Exception as e:
        print(f"Error getting bounds: {e}")
        return None, "Unknown"


def capture_window_with_taskbar():
    """Capture active window plus taskbar"""
    global EXCEL_FILE, TEMP_FOLDER
    try:
        time.sleep(0.3)

        bounds, window_title = get_window_with_taskbar_bounds()

        if bounds is None:
            print("⚠️ Could not get window bounds, falling back to full screen capture")
            screenshot = ImageGrab.grab(all_screens=True)
            capture_type = "Full Screen (Fallback)"
        else:
            screenshot = ImageGrab.grab(bbox=bounds, all_screens=True)
            capture_type = "Window + Taskbar"

        print(f"📸 {capture_type} captured at {datetime.now().strftime('%H:%M:%S')}")

        if not screenshot or not screenshot.getbbox():
            print("❌ Image is black. Ensure Chrome Hardware Acceleration is OFF.")
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_title = "".join(c for c in window_title if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_title = safe_title[:50]
        screenshot_name = os.path.join(TEMP_FOLDER, f"window+taskbar_{safe_title}_{timestamp}.png")
        screenshot.save(screenshot_name)
        print(f"✅ Screenshot saved temporarily: {screenshot_name}")

        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Tax Records"
            headers = ["Timestamp", "Window Title", "Description/Notes", "Screenshot", "Amount", "Category"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col).value = header
                ws.cell(row=1, column=col).font = Font(bold=True)
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.freeze_panes = 'A2'

        next_row = ws.max_row + 1

        if bounds:
            width = bounds[2] - bounds[0]
            height = bounds[3] - bounds[1]
            window_info = f"{window_title} (Window+Taskbar: {width}x{height})"
        else:
            window_info = f"{window_title} (Full Screen)"

        ws.cell(row=next_row, column=1).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=next_row, column=2).value = window_info
        ws.cell(row=next_row, column=3).value = "ENTER NOTES HERE"

        img = Image(screenshot_name)
        max_width = 600
        if img.width > max_width:
            ratio = max_width / img.width
            img.width = max_width
            img.height = int(img.height * ratio)

        ws.add_image(img, f'D{next_row}')
        ws.row_dimensions[next_row].height = max(15, img.height * 0.75)

        validation_exists = False
        if hasattr(ws, 'data_validations'):
            for dv in ws.data_validations.dataValidation:
                if 'F' in str(dv.ranges):
                    validation_exists = True
                    break

        if not validation_exists:
            dv = DataValidation(type="list", formula1='"Income,Expense,Deduction,Asset,Other"')
            dv.add('F2:F1048576')
            ws.add_data_validation(dv)
            print("✅ Category dropdown added to column F")

        wb.save(EXCEL_FILE)
        wb.close()

        if os.path.exists(screenshot_name):
            os.remove(screenshot_name)
            print(f"✅ Temporary file deleted: {screenshot_name}")

        print(f"✅ Excel file saved to: {EXCEL_FILE}")
        print(f"✅ Saved at row {next_row}")
        print(f"📝 Don't forget to fill in your data in columns C, E, and F!")
        print("=" * 50 + "\n")

    except Exception as e:
        print(f"❌ Error: {e}")
        if 'screenshot_name' in locals() and os.path.exists(screenshot_name):
            os.remove(screenshot_name)
            print(f"✅ Cleaned up temporary file")


def create_template():
    """Create a template Excel file with proper formatting"""
    global EXCEL_FILE
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tax Records"
        headers = ["Timestamp", "Window Title", "Description/Notes", "Screenshot", "Amount", "Category"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).value = header
            ws.cell(row=1, column=col).font = Font(bold=True)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.freeze_panes = 'A2'
        dv = DataValidation(type="list", formula1='"Income,Expense,Deduction,Asset,Other"')
        dv.add('F2:F1048576')
        ws.add_data_validation(dv)
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"✅ Template created at: {EXCEL_FILE}")
    except Exception as e:
        print(f"❌ Error creating template: {e}")


# ─── GUI ───────────────────────────────────────────────────────────────────────

class SettingsGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("📸 Tax Screenshot Capture")
        self.root.resizable(False, False)
        self.root.attributes("-topmost", True)

        # Style
        self.root.configure(bg="#f0f4f8")
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TLabel", background="#f0f4f8", font=("Segoe UI", 10))
        style.configure("Header.TLabel", background="#f0f4f8", font=("Segoe UI", 13, "bold"))
        style.configure("Sub.TLabel", background="#f0f4f8", font=("Segoe UI", 9), foreground="#555")
        style.configure("TButton", font=("Segoe UI", 9), padding=4)
        style.configure("Start.TButton", font=("Segoe UI", 11, "bold"), padding=8)
        style.configure("TEntry", font=("Segoe UI", 10))

        self.running = False
        self._build_ui()

    def _build_ui(self):
        root = self.root
        pad = {"padx": 16, "pady": 6}

        # ── Header ──────────────────────────────────────────────────────────
        ttk.Label(root, text="📸 Tax Screenshot Capture Tool", style="Header.TLabel").grid(
            row=0, column=0, columnspan=3, pady=(16, 2), padx=16, sticky="w")
        ttk.Label(root, text="Configure paths then click Start. Use Ctrl+Shift+S to capture.", style="Sub.TLabel").grid(
            row=1, column=0, columnspan=3, padx=16, pady=(0, 10), sticky="w")

        ttk.Separator(root, orient="horizontal").grid(row=2, column=0, columnspan=3, sticky="ew", padx=12)

        # ── Excel File ───────────────────────────────────────────────────────
        ttk.Label(root, text="Excel File (.xlsx):").grid(row=3, column=0, sticky="w", **pad)

        self.excel_var = tk.StringVar(value=DEFAULT_EXCEL_FILE)
        excel_entry = ttk.Entry(root, textvariable=self.excel_var, width=52)
        excel_entry.grid(row=3, column=1, sticky="ew", padx=(0, 4), pady=6)

        ttk.Button(root, text="Browse…", command=self._browse_excel).grid(row=3, column=2, padx=(0, 16), pady=6)

        # ── Temp Folder ──────────────────────────────────────────────────────
        ttk.Label(root, text="Temp Folder:").grid(row=4, column=0, sticky="w", **pad)

        self.temp_var = tk.StringVar(value=DEFAULT_TEMP_FOLDER)
        temp_entry = ttk.Entry(root, textvariable=self.temp_var, width=52)
        temp_entry.grid(row=4, column=1, sticky="ew", padx=(0, 4), pady=6)

        ttk.Button(root, text="Browse…", command=self._browse_temp).grid(row=4, column=2, padx=(0, 16), pady=6)

        ttk.Separator(root, orient="horizontal").grid(row=5, column=0, columnspan=3, sticky="ew", padx=12, pady=(4, 0))

        # ── Status / Log ─────────────────────────────────────────────────────
        ttk.Label(root, text="Log:").grid(row=6, column=0, sticky="nw", padx=16, pady=(8, 0))

        self.log_text = tk.Text(root, height=8, width=62, font=("Consolas", 9),
                                state="disabled", bg="#1e1e1e", fg="#d4d4d4",
                                relief="flat", bd=4)
        self.log_text.grid(row=6, column=1, columnspan=2, padx=(0, 16), pady=(8, 4), sticky="ew")

        # ── Buttons ──────────────────────────────────────────────────────────
        btn_frame = tk.Frame(root, bg="#f0f4f8")
        btn_frame.grid(row=7, column=0, columnspan=3, pady=(4, 16), padx=16, sticky="e")

        self.start_btn = ttk.Button(btn_frame, text="▶  Start Listener", style="Start.TButton",
                                    command=self._start)
        self.start_btn.pack(side="left", padx=(0, 8))

        self.stop_btn = ttk.Button(btn_frame, text="⏹  Stop", command=self._stop, state="disabled")
        self.stop_btn.pack(side="left")

        # Status indicator
        self.status_var = tk.StringVar(value="⚪  Idle")
        ttk.Label(root, textvariable=self.status_var, style="Sub.TLabel").grid(
            row=8, column=0, columnspan=3, padx=16, pady=(0, 12), sticky="w")

        root.columnconfigure(1, weight=1)

    # ── Browse helpers ────────────────────────────────────────────────────────

    def _browse_excel(self):
        path = filedialog.asksaveasfilename(
            title="Select or create Excel file",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=os.path.basename(self.excel_var.get()),
            initialdir=os.path.dirname(self.excel_var.get()) if os.path.dirname(self.excel_var.get()) else "/"
        )
        if path:
            self.excel_var.set(path)

    def _browse_temp(self):
        path = filedialog.askdirectory(
            title="Select Temp Folder for Screenshots",
            initialdir=self.temp_var.get() if os.path.isdir(self.temp_var.get()) else "/"
        )
        if path:
            self.temp_var.set(path)

    # ── Log helper ────────────────────────────────────────────────────────────

    def _log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    # ── Start / Stop ──────────────────────────────────────────────────────────

    def _validate_paths(self):
        excel = self.excel_var.get().strip()
        temp = self.temp_var.get().strip()

        if not excel:
            messagebox.showerror("Missing Path", "Please specify an Excel file path.")
            return False
        if not excel.lower().endswith(".xlsx"):
            messagebox.showerror("Invalid File", "Excel file must end in .xlsx")
            return False

        excel_dir = os.path.dirname(excel)
        if excel_dir and not os.path.isdir(excel_dir):
            try:
                os.makedirs(excel_dir, exist_ok=True)
                self._log(f"📁 Created directory: {excel_dir}")
            except Exception as e:
                messagebox.showerror("Directory Error", f"Cannot create directory:\n{e}")
                return False

        if not temp:
            messagebox.showerror("Missing Path", "Please specify a temp folder.")
            return False
        if not os.path.isdir(temp):
            try:
                os.makedirs(temp, exist_ok=True)
                self._log(f"📁 Created temp folder: {temp}")
            except Exception as e:
                messagebox.showerror("Directory Error", f"Cannot create temp folder:\n{e}")
                return False

        return True

    def _start(self):
        global EXCEL_FILE, TEMP_FOLDER

        if not self._validate_paths():
            return

        EXCEL_FILE = self.excel_var.get().strip()
        TEMP_FOLDER = self.temp_var.get().strip()

        if not os.path.exists(EXCEL_FILE):
            create_template()
            self._log(f"✅ Template Excel created: {EXCEL_FILE}")

        # Register hotkey in a background thread so GUI stays responsive
        keyboard.add_hotkey('ctrl+shift+s', self._capture_and_log)

        self.running = True
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.status_var.set("🟢  Listening — press Ctrl+Shift+S to capture")
        self._log("─" * 50)
        self._log(f"📁 Excel : {EXCEL_FILE}")
        self._log(f"📁 Temp  : {TEMP_FOLDER}")
        self._log("✅ Listener started. Switch to any window and press Ctrl+Shift+S.")

    def _capture_and_log(self):
        """Wrapper so we can log the result back to the GUI"""
        self._log(f"📸 Capturing at {datetime.now().strftime('%H:%M:%S')}…")
        try:
            capture_window_with_taskbar()
            self._log("✅ Capture saved to Excel.")
        except Exception as e:
            self._log(f"❌ Error: {e}")

    def _stop(self):
        keyboard.unhook_all_hotkeys()
        self.running = False
        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.status_var.set("⚪  Idle")
        self._log("⏹ Listener stopped.")

    # ── Run ───────────────────────────────────────────────────────────────────

    def run(self):
        self.root.mainloop()
        # Clean up hotkeys when window is closed
        if self.running:
            keyboard.unhook_all_hotkeys()


# ─── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Check / install required packages
    required = {'pygetwindow': 'pygetwindow', 'PIL': 'pillow', 'keyboard': 'keyboard',
                'pyautogui': 'pyautogui', 'openpyxl': 'openpyxl'}
    for module, pkg in required.items():
        try:
            __import__(module)
        except ImportError:
            print(f"Installing {pkg}…")
            os.system(f"pip install {pkg}")

    app = SettingsGUI()
    app.run()